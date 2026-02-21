"""
Excel报告生成器 - 所有分析数据以规范Excel表格呈现
"""
import io
from datetime import datetime
from typing import List, Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from business_semantic_report_v3 import (
    BusinessSemanticReportGeneratorV3,
    BusinessMetric,
    DimensionAnalysis,
    KeyInsight
)


class ExcelReportGenerator:
    """Excel报告生成器 - 所有数据以规范表格呈现"""
    
    def __init__(self, generator: BusinessSemanticReportGeneratorV3):
        self.generator = generator
        self.wb = Workbook()
        # 删除默认sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
    
    def _create_sheet(self, title: str) -> Any:
        """创建新的工作表"""
        ws = self.wb.create_sheet(title=title)
        return ws
    
    def _format_header(self, cell):
        """格式化表头"""
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _format_data_cell(self, cell, bold: bool = False):
        """格式化数据单元格"""
        cell.font = Font(name='微软雅黑', size=10, bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _add_title(self, ws, title: str, row: int = 1) -> int:
        """添加标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name='微软雅黑', size=14, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        return row + 2
    
    def _add_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int, title: str = "") -> int:
        """将DataFrame添加到工作表"""
        if title:
            # 添加小标题
            cell = ws.cell(row=start_row, column=1, value=title)
            cell.font = Font(name='微软雅黑', size=12, bold=True)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df.columns))
            start_row += 1
        
        # 添加表头
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            self._format_header(cell)
        
        # 添加数据
        for row_idx, row_data in enumerate(df.values, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._format_data_cell(cell)
        
        # 调整列宽
        for col_idx, col_name in enumerate(df.columns, 1):
            max_length = len(str(col_name))
            for row_data in df.values:
                cell_length = len(str(row_data[col_idx - 1]))
                if cell_length > max_length:
                    max_length = cell_length
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = max_length + 4
        
        return start_row + len(df) + 2
    
    def generate_full_report(self, title: str = "业务分析报告") -> io.BytesIO:
        """生成完整的Excel报告"""
        # 执行分析
        self.generator._analyze_business_metrics()
        self.generator._analyze_by_dimensions()
        self.generator._extract_key_insights()
        self.generator._generate_recommendations()
        
        # Sheet 1: 报告概览
        ws_overview = self._create_sheet("报告概览")
        current_row = 1
        
        # 主标题
        cell = ws_overview.cell(row=current_row, column=1, value=title)
        cell.font = Font(name='微软雅黑', size=16, bold=True)
        ws_overview.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 2
        
        # 生成时间
        cell = ws_overview.cell(row=current_row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        cell.font = Font(name='微软雅黑', size=10)
        current_row += 2
        
        # 数据概览
        cell = ws_overview.cell(row=current_row, column=1, value="数据概览")
        cell.font = Font(name='微软雅黑', size=12, bold=True)
        current_row += 1
        
        overview_data = [
            ['项目', '数值', '说明'],
            ['总记录数', len(self.generator.df), '数据总行数'],
            ['数值指标数', len(self.generator.numeric_columns), '可量化的业务指标'],
            ['分类维度数', len(self.generator.categorical_columns), '可用于分组的维度'],
            ['维度分析数', len(self.generator.dimension_analyses), '维度分析结果数量'],
            ['关键洞察数', len(self.generator.key_insights), '提取的关键洞察数量'],
        ]
        
        for row_idx, row_data in enumerate(overview_data, current_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == current_row:
                    self._format_header(cell)
                else:
                    self._format_data_cell(cell)
        
        current_row += len(overview_data) + 2
        
        # Sheet 2: 核心业务指标
        ws_metrics = self._create_sheet("核心业务指标")
        current_row = 1
        
        current_row = self._add_title(ws_metrics, "一、核心业务指标", current_row)
        
        if self.generator.business_metrics:
            metrics_df = self._format_metrics_table(self.generator.business_metrics)
            current_row = self._add_dataframe_to_sheet(ws_metrics, metrics_df, current_row, "业务指标统计表")
            
            # 添加说明
            ws_metrics.cell(row=current_row, column=1, value="指标说明:")
            ws_metrics.cell(row=current_row, column=1).font = Font(name='微软雅黑', size=10, bold=True)
            current_row += 1
            
            explanations = [
                "• 汇总: 该指标所有数据的总和，反映整体规模",
                "• 平均: 该指标的平均值，反映一般水平",
                "• 最大/最小: 该指标的极值，反映波动范围"
            ]
            for exp in explanations:
                ws_metrics.cell(row=current_row, column=1, value=exp)
                ws_metrics.cell(row=current_row, column=1).font = Font(name='微软雅黑', size=9)
                current_row += 1
        
        # Sheet 3: 维度分析
        ws_dim = self._create_sheet("维度分析")
        current_row = 1
        
        current_row = self._add_title(ws_dim, "二、维度分析", current_row)
        
        if self.generator.dimension_analyses:
            for i, analysis in enumerate(self.generator.dimension_analyses, 1):
                # 维度标题
                dim_title = f"{i}. {analysis.dimension_name}维度 - {analysis.metric_name}"
                cell = ws_dim.cell(row=current_row, column=1, value=dim_title)
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                ws_dim.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 1
                
                # 分析总结
                cell = ws_dim.cell(row=current_row, column=1, value=f"分析总结: {analysis.summary}")
                cell.font = Font(name='微软雅黑', size=10, bold=True)
                ws_dim.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 2
                
                # 维度数据表格
                dim_df = self._format_dimension_table(analysis)
                current_row = self._add_dataframe_to_sheet(ws_dim, dim_df, current_row, "详细数据")
                
                # 详细说明
                if analysis.top_values:
                    top1 = analysis.top_values[0]
                    ws_dim.cell(row=current_row, column=1, value="详细说明:")
                    ws_dim.cell(row=current_row, column=1).font = Font(name='微软雅黑', size=10, bold=True)
                    current_row += 1
                    
                    ws_dim.cell(row=current_row, column=1, value=f"• 表现最优: {top1.get(analysis.dimension_name, '未知')}, 汇总值{top1.get('汇总', 0):.2f}")
                    current_row += 1
                    
                    if len(analysis.top_values) > 1:
                        top2 = analysis.top_values[1]
                        ws_dim.cell(row=current_row, column=1, value=f"• 排名第二: {top2.get(analysis.dimension_name, '未知')}, 汇总值{top2.get('汇总', 0):.2f}")
                        current_row += 1
                    
                    ws_dim.cell(row=current_row, column=1, value=f"• 数据覆盖: 共{len(analysis.data)}个类别, 平均每个类别{analysis.data['平均'].mean():.2f}")
                    current_row += 2
        
        # Sheet 4: 关键洞察
        ws_insights = self._create_sheet("关键洞察")
        current_row = 1
        
        current_row = self._add_title(ws_insights, "三、关键洞察", current_row)
        
        if self.generator.key_insights:
            for i, insight in enumerate(self.generator.key_insights, 1):
                type_names = {
                    'anomaly': '异常发现',
                    'opportunity': '业务机会',
                    'pattern': '数据模式',
                    'risk': '潜在风险'
                }
                type_name = type_names.get(insight.insight_type, insight.insight_type)
                
                # 洞察标题
                insight_title = f"{i}. {insight.title} ({type_name})"
                cell = ws_insights.cell(row=current_row, column=1, value=insight_title)
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                ws_insights.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 1
                
                # 洞察描述
                cell = ws_insights.cell(row=current_row, column=1, value=f"描述: {insight.description}")
                cell.font = Font(name='微软雅黑', size=10)
                ws_insights.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 2
                
                # 支撑数据表格
                if insight.supporting_table is not None and not insight.supporting_table.empty:
                    current_row = self._add_dataframe_to_sheet(ws_insights, insight.supporting_table, current_row, "支撑数据详情")
                elif insight.supporting_data:
                    support_df = pd.DataFrame([insight.supporting_data])
                    current_row = self._add_dataframe_to_sheet(ws_insights, support_df, current_row, "支撑数据详情")
                
                # 置信度和优先级
                cell = ws_insights.cell(row=current_row, column=1, value=f"置信度: {insight.confidence} | 优先级: {insight.priority}")
                cell.font = Font(name='微软雅黑', size=9, italic=True)
                current_row += 2
        
        # Sheet 5: 业务建议
        ws_rec = self._create_sheet("业务建议")
        current_row = 1
        
        current_row = self._add_title(ws_rec, "四、业务建议", current_row)
        
        if self.generator.recommendations:
            for i, rec in enumerate(self.generator.recommendations, 1):
                type_names = {
                    'improvement': '改进建议',
                    'opportunity': '机会建议',
                    'optimization': '优化建议'
                }
                type_name = type_names.get(rec.recommendation_type, rec.recommendation_type)
                
                # 建议标题
                rec_title = f"{i}. {rec.title} ({type_name})"
                cell = ws_rec.cell(row=current_row, column=1, value=rec_title)
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                ws_rec.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 1
                
                # 建议内容
                cell = ws_rec.cell(row=current_row, column=1, value=f"内容: {rec.description}")
                cell.font = Font(name='微软雅黑', size=10)
                ws_rec.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                current_row += 1
                
                if rec.expected_impact:
                    cell = ws_rec.cell(row=current_row, column=1, value=f"预期效果: {rec.expected_impact}")
                    cell.font = Font(name='微软雅黑', size=10)
                    ws_rec.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                    current_row += 1
                
                cell = ws_rec.cell(row=current_row, column=1, value=f"实施难度: {rec.implementation_difficulty}")
                cell.font = Font(name='微软雅黑', size=9, italic=True)
                current_row += 2
        
        # 保存到内存
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def _format_metrics_table(self, metrics: List[BusinessMetric]) -> pd.DataFrame:
        """格式化业务指标表"""
        data = []
        for m in metrics:
            data.append({
                '指标名称': m.name,
                '指标值': m.format_str.format(m.value),
                '单位': m.unit,
                '指标说明': m.description
            })
        return pd.DataFrame(data)
    
    def _format_dimension_table(self, analysis: DimensionAnalysis) -> pd.DataFrame:
        """格式化维度分析表"""
        df = analysis.data.copy()
        # 添加排名列
        df['排名'] = range(1, len(df) + 1)
        # 重新排列列顺序
        cols = ['排名', analysis.dimension_name, '汇总', '平均', '计数']
        df = df[[c for c in cols if c in df.columns]]
        return df


def generate_excel_report(df: pd.DataFrame, title: str = "业务分析报告") -> bytes:
    """生成Excel报告的便捷函数"""
    try:
        # 创建分析器
        analyzer = BusinessSemanticReportGeneratorV3(df)
        
        # 创建Excel生成器
        excel_gen = ExcelReportGenerator(analyzer)
        
        # 生成报告
        output = excel_gen.generate_full_report(title)
        
        return output.getvalue()
    except Exception as e:
        print(f"[generate_excel_report] 生成Excel报告失败: {e}")
        import traceback
        traceback.print_exc()
        return None
