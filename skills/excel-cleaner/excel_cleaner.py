"""
Excel Cleaner - Core Module
处理 Excel/CSV 文件的合并单元格、多级表头、关键列填充
"""

import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
import io
import zipfile
import numpy as np
import re


class ExcelCleaner:
    """Excel 数据清洗器"""
    
    def __init__(self, api_key=None):
        """
        初始化清洗器
        
        Args:
            api_key: 保留参数，不再使用（向后兼容）
        """
        self.api_key = api_key

    def get_sheet_names(self, file_content):
        """
        获取文件中的所有工作表名称
        
        Args:
            file_content: 文件路径或文件对象
            
        Returns:
            list: 工作表名称列表
        """
        try:
            # 处理文件路径
            if isinstance(file_content, str):
                with open(file_content, 'rb') as f:
                    return self._get_sheet_names_from_file(f)
            else:
                file_content.seek(0)
                return self._get_sheet_names_from_file(file_content)
        except Exception as e:
            return ['Sheet1']
    
    def _get_sheet_names_from_file(self, file_obj):
        """从文件对象获取工作表名称"""
        file_obj.seek(0)
        
        # 检查文件名
        name = getattr(file_obj, 'name', '')
        
        if name.endswith('.csv'):
            return ['Sheet1']
        
        if name.endswith('.xls'):
            excel_file = pd.ExcelFile(file_obj, engine='xlrd')
            return excel_file.sheet_names
        
        # 默认 .xlsx
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        return wb.sheetnames

    def load_and_fill_merged_cells(self, file_content, sheet_name=None):
        """
        加载文件并处理合并单元格
        
        将合并单元格解合并，并用左上角值填充所有单元格
        
        Args:
            file_content: 文件路径或文件对象
            sheet_name: 工作表名称，默认第一个表
            
        Returns:
            DataFrame: 处理后的数据框
        """
        # 处理文件路径
        if isinstance(file_content, str):
            with open(file_content, 'rb') as f:
                return self._load_and_process(f, sheet_name)
        else:
            file_content.seek(0)
            return self._load_and_process(file_content, sheet_name)
    
    def _load_and_process(self, file_obj, sheet_name):
        """实际加载和处理逻辑"""
        file_obj.seek(0)
        name = getattr(file_obj, 'name', '')
        
        # CSV 格式
        if name.endswith('.csv'):
            df = pd.read_csv(file_obj, header=None)
            return df
        
        # 旧版 Excel (.xls)
        if name.endswith('.xls'):
            df = pd.read_excel(file_obj, header=None, engine='xlrd', sheet_name=sheet_name)
            return df
        
        # 新版 Excel (.xlsx)
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except zipfile.BadZipFile:
            raise ValueError("无效的 Excel 文件格式，请确保是有效的 .xlsx 文件")
        
        # 选择工作表
        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        # 处理合并单元格
        merged_ranges = list(sheet.merged_cells.ranges)
        
        for merged_cell in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            
            # 解合并
            sheet.unmerge_cells(str(merged_cell))
            
            # 填充所有单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value
        
        # 转换为 DataFrame
        df = pd.DataFrame(sheet.values)
        wb.close()
        
        # 垃圾回收
        import gc
        gc.collect()
        
        return df

    def process_headers(self, df, header_rows, separator="_", max_length=100, keep_case=True):
        """
        处理多级表头，合并为单级字段
        
        策略：Dirty Join -> Regex Clean
        
        Args:
            df: 原始数据框
            header_rows: 表头行索引列表（0-based）
            separator: 分隔符，默认 "_"
            max_length: 列名最大长度
            keep_case: 是否保持原大小写
            
        Returns:
            list: 处理后的列名列表
        """
        if not header_rows:
            return [f"Column_{i}" for i in range(df.shape[1])]
        
        # 1. 准备数据
        header_df = df.iloc[header_rows]
        
        # 水平前向填充（处理视觉分组）
        header_df = header_df.replace("", np.nan).ffill(axis=1).fillna("").astype(str)
        
        # 2. 逐行清理
        cleaned_series_list = []
        for idx in range(len(header_df)):
            s = header_df.iloc[idx]
            # 移除控制字符
            s = s.str.replace(r'[\x00-\x1f\x7f]', '', regex=True)
            # 去除空白
            s = s.str.strip()
            # 处理大小写
            if not keep_case:
                s = s.str.lower()
            cleaned_series_list.append(s)
        
        if not cleaned_series_list:
            return []
        
        # 3. 语义去重（相同值设为空字符串）
        for i in range(1, len(cleaned_series_list)):
            curr = cleaned_series_list[i]
            prev = cleaned_series_list[i-1]
            mask = (curr == prev)
            cleaned_series_list[i] = curr.where(~mask, "")
        
        # 4. 向量化连接
        first_row = cleaned_series_list[0]
        rest_rows = cleaned_series_list[1:]
        
        if rest_rows:
            result_series = first_row.str.cat(rest_rows, sep=separator)
        else:
            result_series = first_row
        
        # 5. 正则清理
        if separator:
            sep_esc = re.escape(separator)
            # 合并多个分隔符
            result_series = result_series.str.replace(f"{sep_esc}{{2,}}", separator, regex=True)
            # 去除首尾分隔符
            result_series = result_series.str.strip(separator)
        
        # 6. 最终处理
        result_series = result_series.replace("", "Unnamed")
        
        # 限制长度
        if max_length:
            result_series = result_series.str.slice(0, max_length)
        
        # 7. 去重（添加数字后缀）
        if result_series.duplicated().any():
            counts = result_series.groupby(result_series).cumcount()
            mask = counts > 0
            result_series.loc[mask] = result_series.loc[mask] + "_" + counts.loc[mask].astype(str)
        
        return result_series.tolist()

    def process_key_columns(self, df, key_columns):
        """
        对关键列执行向下填充（forward fill）
        
        用于处理垂直合并单元格的数据丢失问题
        
        Args:
            df: 数据框（已设置列名）
            key_columns: 关键列索引列表（0-based）
            
        Returns:
            DataFrame: 处理后的数据框
        """
        if not key_columns:
            return df
        
        # 创建副本，避免修改原数据
        df = df.copy()
        
        for col_idx in key_columns:
            if not isinstance(col_idx, int):
                continue
            
            if 0 <= col_idx < df.shape[1]:
                col_name = df.columns[col_idx]
                
                # 将空字符串转为 NaN
                df[col_name] = df[col_name].replace(r'^\s*$', np.nan, regex=True)
                
                # 向下填充
                df[col_name] = df[col_name].ffill()
                
                # 填充剩余 NaN 为空字符串
                df[col_name] = df[col_name].fillna("")
        
        return df

    def clean_data(self, file_content, header_rows, data_start_row, key_columns=None, separator="_", sheet_name=None):
        """
        主清洗流程
        
        Args:
            file_content: 文件路径或文件对象
            header_rows: 表头行索引列表（0-based）
            data_start_row: 数据开始行索引（0-based）
            key_columns: 关键列索引列表（0-based），可选
            separator: 多级表头分隔符
            sheet_name: 工作表名称
            
        Returns:
            dict: 包含清洗结果的字典
            {
                "raw_df": DataFrame,      # 原始数据预览（前20行）
                "cleaned_df": DataFrame,  # 清洗后的完整数据
                "structure_info": dict    # 结构信息
            }
        """
        # 1. 加载并处理合并单元格
        raw_df = self.load_and_fill_merged_cells(file_content, sheet_name=sheet_name)
        
        # 2. 处理表头
        new_columns = self.process_headers(raw_df, header_rows, separator=separator)
        
        # 3. 提取数据
        cleaned_df = raw_df.iloc[data_start_row:].copy()
        
        # 处理列数不匹配
        if len(new_columns) == cleaned_df.shape[1]:
            cleaned_df.columns = new_columns
        else:
            print(f"警告：列数不匹配（{len(new_columns)} vs {cleaned_df.shape[1]}），使用默认索引")
            cleaned_df.columns = [f"Col_{i}" for i in range(cleaned_df.shape[1])]
        
        # 4. 清理数据体
        cleaned_df.dropna(how='all', inplace=True)
        cleaned_df.reset_index(drop=True, inplace=True)
        
        # 5. 处理关键列
        if key_columns:
            cleaned_df = self.process_key_columns(cleaned_df, key_columns)
        
        return {
            "raw_df": raw_df.head(20),
            "cleaned_df": cleaned_df,
            "structure_info": {
                "header_rows": header_rows,
                "data_start_row": data_start_row
            }
        }
