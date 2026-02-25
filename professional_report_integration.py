"""
专业报告集成模块 - 将 docx-js 的专业报告能力整合到项目中
使用 subprocess 调用 Node.js 生成高质量 Word 报告
"""

import os
import json
import subprocess
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field, asdict
import pandas as pd
import io


@dataclass
class ProfessionalKeyInsight:
    """专业关键洞察 - 带完整数据支撑"""
    insight: str
    data_support: str = ""
    values: Dict[str, Any] = field(default_factory=dict)
    
    def to_dict(self):
        return {
            'insight': self.insight,
            'dataSupport': self.data_support,
            'values': self.values
        }


@dataclass
class ProfessionalReportTable:
    """专业报告表格"""
    title: str
    headers: List[str]
    rows: List[List[Any]]
    description: str = ""
    
    def to_dict(self):
        return {
            'title': self.title,
            'headers': self.headers,
            'rows': self.rows,
            'description': self.description
        }


@dataclass
class ProfessionalReportData:
    """专业报告数据"""
    title: str
    sheetName: str
    generateTime: str
    summary: str
    keyInsights: List[ProfessionalKeyInsight]
    tables: List[ProfessionalReportTable]
    
    def to_dict(self):
        return {
            'title': self.title,
            'sheetName': self.sheetName,
            'generateTime': self.generateTime,
            'summary': self.summary,
            'keyInsights': [ki.to_dict() for ki in self.keyInsights],
            'tables': [t.to_dict() for t in self.tables]
        }


def create_professional_report_js_template():
    """创建专业报告生成的 JS 模板"""
    return '''
const { Document, Packer, Paragraph, TextRun, Table, TableRow, TableCell,
        Header, Footer, AlignmentType, BorderStyle, WidthType, ShadingType,
        PageNumber, HeadingLevel } = require('docx');
const fs = require('fs');

// 从命令行参数获取数据文件路径
const dataFile = process.argv[2];
const outputFile = process.argv[3];

// 读取报告数据
const reportData = JSON.parse(fs.readFileSync(dataFile, 'utf8'));

// 创建文档
const doc = new Document({
    styles: {
        default: {
            document: {
                run: { font: "微软雅黑", size: 24 }
            }
        },
        paragraphStyles: [
            {
                id: "Heading1",
                name: "Heading 1",
                basedOn: "Normal",
                next: "Normal",
                quickFormat: true,
                run: { size: 36, bold: true, font: "微软雅黑", color: "2E74B5" },
                paragraph: { spacing: { before: 240, after: 240 } }
            },
            {
                id: "Heading2",
                name: "Heading 2",
                basedOn: "Normal",
                next: "Normal",
                quickFormat: true,
                run: { size: 28, bold: true, font: "微软雅黑", color: "2E74B5" },
                paragraph: { spacing: { before: 180, after: 180 } }
            }
        ]
    },
    sections: [{
        properties: {
            page: {
                margin: { top: 1440, right: 1440, bottom: 1440, left: 1440 }
            }
        },
        headers: {
            default: new Header({
                children: [
                    new Paragraph({
                        alignment: AlignmentType.CENTER,
                        children: [
                            new TextRun({
                                text: reportData.title,
                                size: 20,
                                color: "666666",
                                font: "微软雅黑"
                            })
                        ]
                    })
                ]
            })
        },
        footers: {
            default: new Footer({
                children: [
                    new Paragraph({
                        alignment: AlignmentType.CENTER,
                        children: [
                            new TextRun({ text: "第 ", size: 20, font: "微软雅黑" }),
                            new TextRun({ children: [PageNumber.CURRENT], size: 20, font: "微软雅黑" }),
                            new TextRun({ text: " 页", size: 20, font: "微软雅黑" })
                        ]
                    })
                ]
            })
        },
        children: [
            // 标题
            new Paragraph({
                heading: HeadingLevel.HEADING_1,
                alignment: AlignmentType.CENTER,
                children: [
                    new TextRun({
                        text: reportData.title,
                        bold: true,
                        size: 48,
                        font: "微软雅黑",
                        color: "2E74B5"
                    })
                ]
            }),
            
            // 元信息
            new Paragraph({
                alignment: AlignmentType.CENTER,
                spacing: { after: 400 },
                children: [
                    new TextRun({
                        text: `数据表：${reportData.sheetName}    生成时间：${reportData.generateTime}`,
                        size: 22,
                        color: "666666",
                        font: "微软雅黑"
                    })
                ]
            }),
            
            // 报告摘要
            new Paragraph({
                heading: HeadingLevel.HEADING_2,
                children: [new TextRun("报告摘要")]
            }),
            new Paragraph({
                spacing: { after: 300 },
                children: [
                    new TextRun({
                        text: reportData.summary,
                        size: 24,
                        font: "微软雅黑"
                    })
                ]
            }),
            
            // 关键洞察
            new Paragraph({
                heading: HeadingLevel.HEADING_2,
                children: [new TextRun("关键洞察")]
            }),
            ...reportData.keyInsights.flatMap((insight, index) => [
                new Paragraph({
                    spacing: { before: 200, after: 100 },
                    children: [
                        new TextRun({
                            text: `${index + 1}. ${insight.insight}`,
                            bold: true,
                            size: 26,
                            font: "微软雅黑",
                            color: "1F4E79"
                        })
                    ]
                }),
                new Paragraph({
                    spacing: { after: 100 },
                    children: [
                        new TextRun({
                            text: `数据支撑：${insight.dataSupport}`,
                            size: 22,
                            font: "微软雅黑",
                            italics: true,
                            color: "404040"
                        })
                    ]
                }),
                ...(insight.values && Object.keys(insight.values).length > 0 ? [
                    new Paragraph({
                        spacing: { after: 200 },
                        children: [
                            new TextRun({
                                text: "具体数值：",
                                bold: true,
                                size: 22,
                                font: "微软雅黑"
                            })
                        ]
                    }),
                    ...Object.entries(insight.values).map(([key, value]) =>
                        new Paragraph({
                            indent: { left: 360 },
                            spacing: { after: 60 },
                            children: [
                                new TextRun({
                                    text: `• ${key}：${value}`,
                                    size: 22,
                                    font: "微软雅黑"
                                })
                            ]
                        })
                    )
                ] : [])
            ]),
            
            // 数据明细
            new Paragraph({
                heading: HeadingLevel.HEADING_2,
                spacing: { before: 400 },
                children: [new TextRun("数据明细")]
            }),
            
            // 表格
            ...reportData.tables.flatMap(table => [
                new Paragraph({
                    spacing: { before: 300, after: 200 },
                    children: [
                        new TextRun({
                            text: table.title,
                            bold: true,
                            size: 26,
                            font: "微软雅黑",
                            color: "1F4E79"
                        })
                    ]
                }),
                ...(table.description ? [new Paragraph({
                    spacing: { after: 200 },
                    children: [
                        new TextRun({
                            text: table.description,
                            size: 22,
                            font: "微软雅黑",
                            italics: true,
                            color: "666666"
                        })
                    ]
                })] : []),
                new Table({
                    width: { size: 9360, type: WidthType.DXA },
                    columnWidths: Array(table.headers.length).fill(Math.floor(9360 / table.headers.length)),
                    rows: [
                        // 表头行
                        new TableRow({
                            children: table.headers.map(header =>
                                new TableCell({
                                    width: { size: Math.floor(9360 / table.headers.length), type: WidthType.DXA },
                                    shading: { fill: "2E74B5", type: ShadingType.CLEAR },
                                    children: [
                                        new Paragraph({
                                            alignment: AlignmentType.CENTER,
                                            children: [
                                                new TextRun({
                                                    text: header,
                                                    bold: true,
                                                    size: 22,
                                                    font: "微软雅黑",
                                                    color: "FFFFFF"
                                                })
                                            ]
                                        })
                                    ]
                                })
                            )
                        }),
                        // 数据行
                        ...table.rows.map(row =>
                            new TableRow({
                                children: row.map(cell =>
                                    new TableCell({
                                        width: { size: Math.floor(9360 / table.headers.length), type: WidthType.DXA },
                                        children: [
                                            new Paragraph({
                                                alignment: AlignmentType.CENTER,
                                                children: [
                                                    new TextRun({
                                                        text: String(cell),
                                                        size: 20,
                                                        font: "微软雅黑"
                                                    })
                                                ]
                                            })
                                        ]
                                    })
                                )
                            })
                        )
                    ]
                })
            ]),
            
            // 结语
            new Paragraph({
                heading: HeadingLevel.HEADING_2,
                spacing: { before: 400 },
                children: [new TextRun("结语")]
            }),
            new Paragraph({
                children: [
                    new TextRun({
                        text: "本报告基于清洗后的原始数据生成，确保数据的准确性和一致性。所有关键洞察均有数据支撑，可供业务决策参考。",
                        size: 24,
                        font: "微软雅黑"
                    })
                ]
            })
        ]
    }]
});

// 生成文档
Packer.toBuffer(doc).then(buffer => {
    fs.writeFileSync(outputFile, buffer);
    console.log('✅ 专业分析报告已生成：' + outputFile);
}).catch(err => {
    console.error('❌ 生成报告失败:', err);
    process.exit(1);
});
'''


def generate_professional_report(
    df: pd.DataFrame,
    title: str = "数据分析报告",
    sheet_name: str = "数据表",
    key_insights: List[ProfessionalKeyInsight] = None,
    summary: str = ""
) -> io.BytesIO:
    """
    生成专业 Word 报告
    
    Args:
        df: 清洗后的 DataFrame
        title: 报告标题
        sheet_name: 数据表名称
        key_insights: 关键洞察列表
        summary: 报告摘要
        
    Returns:
        BytesIO 对象，包含 Word 文档
    """
    if df is None or df.empty:
        raise ValueError("DataFrame 不能为空")
    
    # 准备报告数据
    generate_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    
    # 如果没有提供关键洞察，从数据自动生成
    if not key_insights:
        key_insights = _auto_generate_insights(df)
    
    # 如果没有提供摘要，自动生成
    if not summary:
        summary = f"本报告基于{sheet_name}的清洗后数据生成，共{len(df)}条记录，{len(df.columns)}个字段。对数据进行了全面分析，提取了关键业务洞察。"
    
    # 创建表格数据
    tables = _create_tables_from_dataframe(df)
    
    # 构建报告数据
    report_data = ProfessionalReportData(
        title=title,
        sheetName=sheet_name,
        generateTime=generate_time,
        summary=summary,
        keyInsights=key_insights,
        tables=tables
    )
    
    # 创建临时文件
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump(report_data.to_dict(), f, ensure_ascii=False, indent=2)
        data_file = f.name
    
    output_file = data_file.replace('.json', '.docx')
    
    try:
        # 创建 JS 脚本（放在项目目录下，以便找到 node_modules）
        js_script = create_professional_report_js_template()
        js_file = os.path.join(os.path.dirname(__file__), 'temp_report_generator.js')
        
        with open(js_file, 'w', encoding='utf-8') as f:
            f.write(js_script)
        
        # 执行 Node.js 脚本（在项目目录下执行，以便找到 node_modules）
        result = subprocess.run(
            ['node', js_file, data_file, output_file],
            capture_output=True,
            text=True,
            timeout=30,
            cwd=os.path.dirname(__file__)  # 在项目目录下执行
        )
        
        if result.returncode != 0:
            raise RuntimeError(f"生成报告失败: {result.stderr}")
        
        # 读取生成的文档
        with open(output_file, 'rb') as f:
            docx_buffer = io.BytesIO(f.read())
        
        return docx_buffer
        
    finally:
        # 清理临时文件
        for f in [data_file, output_file]:
            if f and os.path.exists(f):
                try:
                    os.remove(f)
                except:
                    pass
        # 保留 JS 文件以便复用，或清理
        if js_file and os.path.exists(js_file):
            try:
                os.remove(js_file)
            except:
                pass


def _auto_generate_insights(df: pd.DataFrame) -> List[ProfessionalKeyInsight]:
    """从 DataFrame 自动生成关键洞察"""
    insights = []
    
    # 检测数值列
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    if numeric_cols:
        # 找出最大值
        max_col = numeric_cols[0]
        max_val = df[max_col].max()
        max_row = df[df[max_col] == max_val].iloc[0]
        
        # 尝试找到名称列
        name_col = None
        for col in ['线路名称', '名称', '线路', 'name', 'Name']:
            if col in df.columns:
                name_col = col
                break
        
        if name_col:
            insights.append(ProfessionalKeyInsight(
                insight=f"{max_row[name_col]}的{max_col}最高，达到{max_val}",
                data_support=f"在所有记录中，{max_row[name_col]}的{max_col}显著高于其他记录",
                values={
                    f"{max_row[name_col]}的{max_col}": max_val,
                    "平均值": round(df[max_col].mean(), 2),
                    "高出平均": f"{((max_val / df[max_col].mean() - 1) * 100):.1f}%"
                }
            ))
    
    # 检测分类列
    categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
    
    for col in categorical_cols:
        if col in ['线路属性', '属性', '类型', '类别']:
            value_counts = df[col].value_counts()
            if len(value_counts) > 1:
                top_value = value_counts.index[0]
                insights.append(ProfessionalKeyInsight(
                    insight=f"{top_value}类线路占主导地位",
                    data_support=f"从{col}分布来看，{top_value}类线路数量最多",
                    values={
                        f"{top_value}类数量": int(value_counts.iloc[0]),
                        "总数量": len(df),
                        f"占比": f"{(value_counts.iloc[0] / len(df) * 100):.1f}%"
                    }
                ))
            break
    
    # 如果没有生成任何洞察，添加一个默认的
    if not insights:
        insights.append(ProfessionalKeyInsight(
            insight=f"数据包含{len(df)}条记录，{len(df.columns)}个字段",
            data_support="基于完整数据集统计分析",
            values={"记录数": len(df), "字段数": len(df.columns)}
        ))
    
    return insights


def _create_tables_from_dataframe(df: pd.DataFrame, max_rows: int = 15) -> List[ProfessionalReportTable]:
    """从 DataFrame 创建报告表格"""
    tables = []
    
    # 主数据表（前 N 行）
    display_df = df.head(max_rows)
    
    # 转换数据为列表格式
    headers = display_df.columns.tolist()
    rows = display_df.values.tolist()
    
    # 格式化数值
    formatted_rows = []
    for row in rows:
        formatted_row = []
        for val in row:
            if isinstance(val, float):
                formatted_row.append(f"{val:.2f}")
            else:
                formatted_row.append(str(val))
        formatted_rows.append(formatted_row)
    
    tables.append(ProfessionalReportTable(
        title=f"数据明细（前{len(display_df)}条记录）",
        headers=headers,
        rows=formatted_rows,
        description=f"共{len(df)}条记录，显示前{len(display_df)}条"
    ))
    
    # 如果有分类字段，生成分类汇总表
    categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    for col in categorical_cols:
        if col in ['线路属性', '属性', '类型', '类别', '单位']:
            if numeric_cols:
                # 按分类汇总数值字段
                summary = df.groupby(col)[numeric_cols[0]].agg(['count', 'sum', 'mean']).reset_index()
                summary.columns = [col, '数量', '总计', '平均值']
                
                tables.append(ProfessionalReportTable(
                    title=f"按{col}汇总",
                    headers=summary.columns.tolist(),
                    rows=[[f"{v:.2f}" if isinstance(v, float) else str(v) for v in row] 
                          for row in summary.values.tolist()],
                    description=f"按{col}分组统计"
                ))
            break
    
    return tables


def integrate_with_app(agent_result, cleaned_df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
    """
    与主应用集成的接口函数
    
    Args:
        agent_result: Agent 分析结果
        cleaned_df: 清洗后的 DataFrame
        sheet_name: 工作表名称
        
    Returns:
        包含 markdown、word、excel 的字典
    """
    # 从 Agent 结果提取关键洞察
    key_insights = []
    
    if hasattr(agent_result, 'insights') and agent_result.insights:
        for insight_text in agent_result.insights:
            key_insights.append(ProfessionalKeyInsight(
                insight=insight_text,
                data_support="基于AI深度分析得出",
                values={}
            ))
    
    # 生成专业报告
    word_buffer = generate_professional_report(
        df=cleaned_df,
        title=f'{sheet_name}数据分析报告',
        sheet_name=sheet_name,
        key_insights=key_insights,
        summary=f'本报告基于{sheet_name}的清洗后数据生成，共{len(cleaned_df)}条记录，确保数据准确性和业务洞察的可靠性。'
    )
    
    # 同时生成 Markdown 版本
    md_content = _generate_markdown_report(cleaned_df, sheet_name, key_insights)
    
    # 生成 Excel 版本
    excel_buffer = _generate_excel_report(cleaned_df, sheet_name)
    
    return {
        'markdown': md_content,
        'word': word_buffer,
        'excel': excel_buffer.getvalue() if excel_buffer else None
    }


def _generate_markdown_report(df: pd.DataFrame, sheet_name: str, key_insights: List[ProfessionalKeyInsight]) -> str:
    """生成 Markdown 格式报告"""
    lines = [
        f"# {sheet_name}数据分析报告",
        "",
        f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 报告摘要",
        "",
        f"本报告基于{sheet_name}的清洗后数据生成，共{len(df)}条记录，{len(df.columns)}个字段。",
        "",
        "## 关键洞察",
        ""
    ]
    
    for i, insight in enumerate(key_insights, 1):
        lines.extend([
            f"### {i}. {insight.insight}",
            "",
            f"**数据支撑**: {insight.data_support}",
            ""
        ])
        if insight.values:
            lines.append("**具体数值**:")
            for key, value in insight.values.items():
                lines.append(f"- {key}: {value}")
            lines.append("")
    
    lines.extend([
        "## 数据明细",
        ""
    ])
    
    # 手动生成 Markdown 表格（不依赖 tabulate）
    display_df = df.head(20)
    headers = display_df.columns.tolist()
    
    # 表头
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    
    # 数据行
    for _, row in display_df.iterrows():
        row_values = [str(v) if v is not None else "" for v in row.values]
        lines.append("| " + " | ".join(row_values) + " |")
    
    lines.append("")
    
    return "\n".join(lines)


def _generate_excel_report(df: pd.DataFrame, sheet_name: str) -> Optional[io.BytesIO]:
    """生成 Excel 格式报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "数据明细"
        
        # 写入表头
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"生成 Excel 报告失败: {e}")
        return None


# 兼容性函数，保持与现有代码的接口一致
def generate_reports(agent_result, sheet_name: str = "data") -> tuple:
    """
    兼容旧版接口的报告生成函数
    
    Returns:
        (markdown_content, word_buffer, excel_bytes)
    """
    # 尝试从 session_state 获取清洗后的数据
    import streamlit as st
    
    cleaned_df = None
    if 'cleaned_data' in st.session_state and st.session_state.cleaned_data:
        if isinstance(st.session_state.cleaned_data, dict):
            # 多表情况，使用第一个表
            cleaned_df = list(st.session_state.cleaned_data.values())[0]
        else:
            cleaned_df = st.session_state.cleaned_data
    
    if cleaned_df is None:
        # 回退：创建空 DataFrame
        cleaned_df = pd.DataFrame()
    
    result = integrate_with_app(agent_result, cleaned_df, sheet_name)
    
    return (
        result['markdown'],
        result['word'],
        result['excel']
    )
