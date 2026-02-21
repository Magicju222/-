import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
import json
import io
import zipfile
import numpy as np
import datetime
import os

class ExcelCleaner:
    def __init__(self, api_key=None):
        # API Key is no longer needed but kept for signature compatibility
        self.api_key = api_key
        # 大文件阈值（50MB）
        self.large_file_threshold = 50 * 1024 * 1024
        # CSV分块大小
        self.csv_chunksize = 100000

    def get_sheet_names(self, file_content):
        """
        Get list of sheet names from the file.
        """
        try:
            file_content.seek(0)
            file_name = getattr(file_content, 'name', '')
            
            if file_name.endswith('.csv'):
                return ['Sheet1']
            
            if file_name.endswith('.xls'):
                excel_file = pd.ExcelFile(file_content, engine='xlrd')
                return excel_file.sheet_names
            
            # Default .xlsx
            # Use read_only=True for speed when just getting sheet names
            wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            return wb.sheetnames
        except Exception as e:
            # Fallback
            return ['Sheet1']

    def load_and_fill_merged_cells(self, file_content, sheet_name=None):
        """
        Load Excel/CSV, handle merged cells by unmerging and filling values (forward/right fill logic).
        Returns a pandas DataFrame.
        """
        try:
            # Ensure pointer is at start
            file_content.seek(0)
            
            # Check file type - handle both file objects and BytesIO
            file_name = getattr(file_content, 'name', '')
            
            if file_name.endswith('.csv'):
                # 检查文件大小，大文件使用分块读取
                file_content.seek(0, os.SEEK_END)
                file_size = file_content.tell()
                file_content.seek(0)
                
                if file_size > self.large_file_threshold:
                    # 大文件分块读取
                    chunks = []
                    for chunk in pd.read_csv(file_content, header=None, chunksize=self.csv_chunksize):
                        chunks.append(chunk)
                        # 限制内存使用，只保留前100万行
                        if sum(len(c) for c in chunks) >= 1000000:
                            break
                    df = pd.concat(chunks, ignore_index=True)
                    # 如果数据被截断，添加提示
                    if len(df) >= 1000000:
                        print("警告：CSV文件过大，只加载了前100万行数据")
                else:
                    df = pd.read_csv(file_content, header=None)
                return df
                
            if file_name.endswith('.xls'):
                # Legacy Excel format
                # xlrd engine is needed.
                # Note: We cannot easily "unmerge and fill" without openpyxl support.
                # We return standard pandas read (merged cells = NaNs)
                df = pd.read_excel(file_content, header=None, engine='xlrd', sheet_name=sheet_name)
                return df
                
            # Assume .xlsx (default)
            wb = openpyxl.load_workbook(file_content, data_only=True)
        except zipfile.BadZipFile:
            raise ValueError("Invalid Excel file format. Please ensure it is a valid .xlsx file.")
        except Exception as e:
            # Fallback for csv/xls if name check failed or other error
            try:
                file_content.seek(0)
                if str(e).find("Excel file format") != -1: # Maybe it's csv?
                    df = pd.read_csv(file_content, header=None)
                    return df
            except:
                pass
            raise ValueError(f"Failed to load file: {str(e)}")

        # Select specific sheet if provided, otherwise active
        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active

        # Handle merged cells: unmerge and fill with the top-left value
        merged_ranges = list(sheet.merged_cells.ranges)
        
        # We need to sort merged ranges to handle nested merges if any (though rare in Excel structure)
        # But critical: We must fill the cells in the sheet object before converting to values
        for merged_cell in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell))
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            
            sheet.unmerge_cells(str(merged_cell))
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

        # Convert to DataFrame
        # Use generator to save memory, then close workbook immediately
        df = pd.DataFrame(sheet.values)
        wb.close()
        
        # Explicit garbage collection for large objects
        import gc
        gc.collect()
        
        return df

    def process_headers(self, df, header_rows, separator="_", max_length=100, keep_case=True):
        """
        Refactored header processing with advanced cleaning and deduplication.
        OPTIMIZED: Uses Pandas vectorized operations (str.cat) with regex cleanup for speed.
        Strategy: Dirty Join -> Regex Clean.
        """
        if not header_rows:
            return [f"Column_{i}" for i in range(df.shape[1])]
            
        # 1. Prepare Data
        # Use standard object dtype which Pandas optimizes well for str.cat
        header_df = df.iloc[header_rows]
        
        # Horizontal Forward Fill for grouped headers (visual grouping without merge)
        # This ensures that if "Region" is in Col A, and Col B is empty (visually under Region), it inherits "Region".
        header_df = header_df.replace("", np.nan).ffill(axis=1).fillna("").astype(str)
        
        # 2. Row-wise Cleaning
        # Iterate over rows (levels), update in place or list
        cleaned_series_list = []
        for idx in range(len(header_df)):
            s = header_df.iloc[idx]
            # Remove control characters
            s = s.str.replace(r'[\x00-\x1f\x7f]', '', regex=True)
            # Strip whitespace
            s = s.str.strip()
            # Handle case
            if not keep_case:
                s = s.str.lower()
            cleaned_series_list.append(s)
            
        if not cleaned_series_list:
            return []
            
        # 3. Semantic Deduplication (Masking)
        # If value equals previous row value, set to empty string (to be collapsed later)
        # This handles the A -> A -> B case (becomes A -> "" -> B)
        # BUT we must handle the accumulated context. 
        # Actually, if we just set to "", joining "A" and "" with "_" gives "A_".
        # We handle that with regex cleanup later.
        for i in range(1, len(cleaned_series_list)):
            curr = cleaned_series_list[i]
            prev = cleaned_series_list[i-1]
            mask = (curr == prev)
            # Set duplicates to empty string
            cleaned_series_list[i] = curr.where(~mask, "")

        # 4. Vectorized Join (The "Dirty Join" Strategy)
        first_row = cleaned_series_list[0]
        rest_rows = cleaned_series_list[1:]
        
        if rest_rows:
            # Join all series with separator
            result_series = first_row.str.cat(rest_rows, sep=separator)
        else:
            result_series = first_row
            
        # 5. Fast Cleanup (Regex)
        # Collapse multiple separators (e.g. "__") to single ("_")
        # And strip leading/trailing separators (caused by empty strings at start/end)
        if separator:
            import re
            sep_esc = re.escape(separator)
            # Replace 2 or more separators with 1
            result_series = result_series.str.replace(f"{sep_esc}{{2,}}", separator, regex=True)
            # Strip separators from edges
            result_series = result_series.str.strip(separator)

        # 6. Final Processing
        # Handle empty results (all empty strings)
        result_series = result_series.replace("", "Unnamed")
        
        # Enforce max_length
        if max_length:
             result_series = result_series.str.slice(0, max_length)
             
        # 7. Final Deduplication (Uniqueness)
        if result_series.duplicated().any():
            counts = result_series.groupby(result_series).cumcount()
            mask = counts > 0
            # Add suffix only to duplicates
            result_series.loc[mask] = result_series.loc[mask] + "_" + counts.loc[mask].astype(str)
            
        return result_series.tolist()

    def process_key_columns(self, df, key_columns):
        """
        Performs vertical forward-fill on selected key columns.
        This handles 'Index Columns' where cells are merged vertically (e.g. Region A spans 3 rows).
        """
        if not key_columns:
            return df
            
        # Ensure key_columns are valid column names or indices
        # Since we use indices from UI, we need to map them to actual column names
        # BUT wait, the input `df` here is `cleaned_df` which already has new column names.
        # The `key_columns` from UI are integer indices (0-based column index).
        
        # We need to be careful: key_columns indices refer to the position in the FINAL dataframe.
        # So we can use iloc.
        
        for col_idx in key_columns:
            # Type check to avoid TypeError: '<=' not supported between instances of 'int' and 'str'
            if not isinstance(col_idx, int):
                continue
                
            if 0 <= col_idx < df.shape[1]:
                # Convert empty strings to NaN for ffill to work
                # Note: We must be careful not to overwrite legitimate empty strings if any, 
                # but usually in key columns (like Region), empty means "same as above".
                
                # Get column name by index
                col_name = df.columns[col_idx]
                
                # Replace empty strings and None with NaN
                df[col_name] = df[col_name].replace(r'^\s*$', np.nan, regex=True).fillna(np.nan)
                
                # Forward fill
                df[col_name] = df[col_name].ffill()
                
                # Fill remaining NaNs with empty string if desired, or keep as is?
                # Usually better to keep as is or fill with ""
                df[col_name] = df[col_name].fillna("")
                
        return df

    def clean_data(self, file_content, header_rows, data_start_row, key_columns=None, separator="_", sheet_name=None, data_end_row=None, data_end_col=None):
        """
        Main execution pipeline. Strictly uses provided manual structure.
        
        Args:
            file_content: 文件内容
            header_rows: 表头行索引列表（0-based）
            data_start_row: 数据开始行索引（0-based）
            key_columns: 关键列索引列表（0-based），用于向下填充
            separator: 多级表头分隔符
            sheet_name: 工作表名称
            data_end_row: 数据结束行索引（0-based），可选，None表示到最后一行
            data_end_col: 数据结束列索引（0-based），可选，None表示到最后一列
        """
        # 1. Load and fill merged cells
        raw_df = self.load_and_fill_merged_cells(file_content, sheet_name=sheet_name)
        
        # 2. Process Headers
        new_columns = self.process_headers(raw_df, header_rows, separator=separator)
        
        # 3. Extract Data with range constraints
        # Slice from data_start_row to data_end_row (if specified)
        if data_end_row is not None:
            cleaned_df = raw_df.iloc[data_start_row:data_end_row+1].copy()
        else:
            cleaned_df = raw_df.iloc[data_start_row:].copy()
        
        # Apply column range constraint (if specified)
        if data_end_col is not None:
            # Slice columns from 0 to data_end_col (inclusive)
            cleaned_df = cleaned_df.iloc[:, :data_end_col+1].copy()
            # Also truncate new_columns to match
            new_columns = new_columns[:data_end_col+1]
        
        # Handle columns mismatch (rare but possible if header processing is weird)
        if len(new_columns) == cleaned_df.shape[1]:
            cleaned_df.columns = new_columns
        else:
            # Fallback if dimensions don't match
            print("Warning: Column count mismatch. Using default indexing.")
            cleaned_df.columns = [f"Col_{i}" for i in range(cleaned_df.shape[1])]
            
        # 4. Clean Data Body
        # Drop rows that are completely empty
        cleaned_df.dropna(how='all', inplace=True)
        # Reset index
        cleaned_df.reset_index(drop=True, inplace=True)
        
        # 5. Process Key Columns (Vertical Fill)
        if key_columns:
            cleaned_df = self.process_key_columns(cleaned_df, key_columns)
        
        return {
            "raw_df": raw_df.head(20), # Return preview of raw
            "cleaned_df": cleaned_df,
            "structure_info": {
                "header_rows": header_rows,
                "data_start_row": data_start_row,
                "data_end_row": data_end_row,
                "data_end_col": data_end_col
            }
        }
