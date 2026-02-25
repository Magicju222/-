"""
层级化报告生成器
按照系统化层级结构构建分析报告：
1. 总体核心指标（宏观）
2. 指标层级分解（中观）
3. 关联性分析（微观）
每个维度都包含对应的表格数据支撑
"""

import os
import json
import subprocess
import tempfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field
import pandas as pd
import numpy as np
import io


@dataclass
class HierarchicalInsight:
    """层级化洞察"""
    level: str  # '宏观', '中观', '微观'
    dimension: str  # 维度名称
    title: str
    insight: str
    data_support: str
    values: Dict[str, Any] = field(default_factory=dict)
    table_refs: List[str] = field(default_factory=list)  # 关联的表格引用
    recommendation: str = ""


@dataclass
class ReportTable:
    """报告表格"""
    id: str  # 表格唯一标识
    title: str
    headers: List[str]
    rows: List[List[Any]]
    description: str = ""
    metrics: Dict[str, Any] = field(default_factory=dict)  # 指标元数据


class HierarchicalAnalyzer:
    """层级化分析器"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        self.categorical_cols = df.select_dtypes(include=['object']).columns.tolist()
        
        # 识别关键列
        self.name_col = self._detect_name_column()
        self.category_col = self._detect_category_column()
        self.time_col = self._detect_time_column()
        
    def _detect_name_column(self) -> Optional[str]:
        """检测名称列"""
        for col in ['线路名称', '单位名称', '名称', '线路', 'name', 'Name', '部门']:
            if col in self.df.columns:
                return col
        return self.categorical_cols[0] if self.categorical_cols else None
    
    def _detect_category_column(self) -> Optional[str]:
        """检测分类列"""
        for col in ['线路属性', '属性', '类型', '类别', '单位']:
            if col in self.df.columns:
                return col
        # 找唯一值较少的列
        for col in self.categorical_cols:
            if self.df[col].nunique() <= 10:
                return col
        return None
    
    def _detect_time_column(self) -> Optional[str]:
        """检测时间列"""
        for col in ['日期', '时间', '月份', '年份', 'date', 'time']:
            if col in self.df.columns:
                return col
        return None
    
    def analyze_macro_level(self) -> Tuple[List[HierarchicalInsight], List[ReportTable]]:
        """
        宏观层分析 - 总体核心指标
        """
        insights = []
        tables = []
        
        # 洞察1: 总体规模指标
        total_records = len(self.df)
        total_numeric_cols = len(self.numeric_cols)
        
        # 创建总体概况表
        overview_data = []
        for col in self.numeric_cols[:5]:
            overview_data.append([
                col,
                f"{self.df[col].sum():.2f}",
                f"{self.df[col].mean():.2f}",
                f"{self.df[col].max():.2f}",
                f"{self.df[col].min():.2f}"
            ])
        
        overview_table = ReportTable(
            id="T1-总体核心指标",
            title="总体核心指标汇总",
            headers=["指标名称", "总和", "平均值", "最大值", "最小值"],
            rows=overview_data,
            description="数据集总体规模与核心指标统计",
            metrics={"total_records": total_records, "total_fields": len(self.df.columns)}
        )
        tables.append(overview_table)
        
        insights.append(HierarchicalInsight(
            level="宏观",
            dimension="总体规模",
            title="数据总体概况",
            insight=f"数据集包含{total_records}条记录，涵盖{len(self.df.columns)}个字段，其中数值指标{total_numeric_cols}个",
            data_support="基于完整数据集统计",
            values={
                "记录总数": total_records,
                "字段总数": len(self.df.columns),
                "数值指标数": total_numeric_cols,
                "分类维度数": len(self.categorical_cols)
            },
            table_refs=["T1-总体核心指标"],
            recommendation="建议持续监控数据规模变化，确保数据完整性"
        ))
        
        # 洞察2: 核心指标Top排名（宏观视角）
        if self.numeric_cols and self.name_col:
            primary_metric = self.numeric_cols[0]
            top5 = self.df.nlargest(5, primary_metric)
            
            top_data = []
            for idx, (_, row) in enumerate(top5.iterrows(), 1):
                top_data.append([
                    idx,
                    row[self.name_col],
                    f"{row[primary_metric]:.2f}",
                    f"{(row[primary_metric] / self.df[primary_metric].sum() * 100):.1f}%"
                ])
            
            top_table = ReportTable(
                id="T2-宏观Top排名",
                title=f"{primary_metric} Top 5 排名",
                headers=["排名", self.name_col, primary_metric, "占比"],
                rows=top_data,
                description=f"按{primary_metric}排序的前5名",
                metrics={"total": self.df[primary_metric].sum(), "average": self.df[primary_metric].mean()}
            )
            tables.append(top_table)
            
            top1_name = top5.iloc[0][self.name_col]
            top1_value = top5.iloc[0][primary_metric]
            top1_share = (top1_value / self.df[primary_metric].sum() * 100)
            
            # 计算前5名合计占比
            top5_total = sum([float(row[2]) for row in top_data])
            top5_share = (top5_total / self.df[primary_metric].sum() * 100)
            
            insights.append(HierarchicalInsight(
                level="宏观",
                dimension="核心指标",
                title=f"{primary_metric}领先者分析",
                insight=f"{top1_name}的{primary_metric}最高，达到{top1_value:.2f}，占总体的{top1_share:.1f}%",
                data_support=f"前5名合计占比{top5_share:.1f}%",
                values={
                    "第一名": top1_name,
                    "指标值": f"{top1_value:.2f}",
                    "占比": f"{top1_share:.1f}%",
                    "前5名合计": f"{top5_total:.2f}"
                },
                table_refs=["T2-宏观Top排名"],
                recommendation=f"建议分析{top1_name}的成功经验，制定推广策略"
            ))
        
        return insights, tables
    
    def analyze_meso_level(self) -> Tuple[List[HierarchicalInsight], List[ReportTable]]:
        """
        中观层分析 - 指标层级分解（从大到小）
        """
        insights = []
        tables = []
        
        # 洞察3: 分类维度分析（部门/类型级别）
        if self.category_col and self.numeric_cols:
            primary_metric = self.numeric_cols[0]
            
            # 按分类汇总
            category_summary = self.df.groupby(self.category_col).agg({
                primary_metric: ['sum', 'mean', 'count'],
                self.name_col: 'count'
            }).reset_index()
            
            category_summary.columns = [self.category_col, '总计', '平均值', '记录数', '项目数']
            category_summary = category_summary.sort_values('总计', ascending=False)
            
            # 计算占比
            total_sum = category_summary['总计'].sum()
            category_summary['占比'] = category_summary['总计'].apply(lambda x: f"{(x/total_sum*100):.1f}%")
            
            cat_data = []
            for _, row in category_summary.iterrows():
                cat_data.append([
                    row[self.category_col],
                    f"{row['总计']:.2f}",
                    f"{row['平均值']:.2f}",
                    int(row['记录数']),
                    row['占比']
                ])
            
            cat_table = ReportTable(
                id="T3-分类维度分析",
                title=f"按{self.category_col}分解分析",
                headers=[self.category_col, "总计", "平均值", "记录数", "占比"],
                rows=cat_data,
                description=f"按{self.category_col}维度分解{primary_metric}指标",
                metrics={"categories": len(cat_data)}
            )
            tables.append(cat_table)
            
            # 找出主导类别
            dominant = category_summary.iloc[0]
            insights.append(HierarchicalInsight(
                level="中观",
                dimension="分类维度",
                title=f"{self.category_col}分布特征",
                insight=f"{dominant[self.category_col]}类{primary_metric}最高，达到{dominant['总计']:.2f}，占比{dominant['占比']}",
                data_support=f"共{len(cat_data)}个类别，分布不均",
                values={
                    "主导类别": dominant[self.category_col],
                    "指标值": f"{dominant['总计']:.2f}",
                    "占比": dominant['占比'],
                    "类别数量": len(cat_data),
                    "平均每个类别": f"{total_sum/len(cat_data):.2f}"
                },
                table_refs=["T3-分类维度分析"],
                recommendation=f"建议优化{dominant[self.category_col]}类的资源配置，同时关注弱势类别"
            ))
            
            # 洞察4: 多指标对比分析（同一类别下多个指标）
            if len(self.numeric_cols) >= 2:
                multi_metric_data = []
                for cat in category_summary[self.category_col].head(5):
                    cat_data_row = [cat]
                    for metric in self.numeric_cols[:4]:
                        value = self.df[self.df[self.category_col] == cat][metric].sum()
                        cat_data_row.append(f"{value:.2f}")
                    multi_metric_data.append(cat_data_row)
                
                multi_table = ReportTable(
                    id="T4-多指标对比",
                    title=f"各类别多指标对比",
                    headers=[self.category_col] + self.numeric_cols[:4],
                    rows=multi_metric_data,
                    description="同一类别下不同指标的数值对比",
                    metrics={"metrics_compared": len(self.numeric_cols[:4])}
                )
                tables.append(multi_table)
                
                # 计算指标间差异
                first_cat = category_summary.iloc[0][self.category_col]
                second_cat = category_summary.iloc[1][self.category_col] if len(category_summary) > 1 else None
                
                if second_cat:
                    first_value = category_summary.iloc[0]['总计']
                    second_value = category_summary.iloc[1]['总计']
                    gap = ((first_value - second_value) / second_value * 100)
                    
                    insights.append(HierarchicalInsight(
                        level="中观",
                        dimension="指标对比",
                        title="类别间差距分析",
                        insight=f"{first_cat}与{second_cat}的{primary_metric}差距为{gap:.1f}%",
                        data_support="基于前两名对比分析",
                        values={
                            "第一名": first_cat,
                            "第二名": second_cat,
                            "差距百分比": f"{gap:.1f}%",
                            "第一名数值": f"{first_value:.2f}",
                            "第二名数值": f"{second_value:.2f}"
                        },
                        table_refs=["T3-分类维度分析", "T4-多指标对比"],
                        recommendation="建议分析差距原因，制定缩小差距的措施"
                    ))
        
        return insights, tables
    
    def analyze_micro_level(self) -> Tuple[List[HierarchicalInsight], List[ReportTable]]:
        """
        微观层分析 - 关联性分析
        """
        insights = []
        tables = []
        
        # 洞察5: 指标相关性分析
        if len(self.numeric_cols) >= 2:
            corr_matrix = self.df[self.numeric_cols].corr()
            
            # 找出强相关性
            correlations = []
            for i in range(len(self.numeric_cols)):
                for j in range(i+1, len(self.numeric_cols)):
                    corr_val = corr_matrix.iloc[i, j]
                    if abs(corr_val) > 0.3:  # 中等以上相关
                        correlations.append({
                            'col1': self.numeric_cols[i],
                            'col2': self.numeric_cols[j],
                            'corr': corr_val
                        })
            
            if correlations:
                # 按相关性排序
                correlations.sort(key=lambda x: abs(x['corr']), reverse=True)
                
                corr_data = []
                for corr in correlations[:6]:
                    strength = '强正相关' if corr['corr'] > 0.7 else (
                        '中等正相关' if corr['corr'] > 0.3 else (
                            '强负相关' if corr['corr'] < -0.7 else '中等负相关'
                        )
                    )
                    corr_data.append([
                        corr['col1'],
                        corr['col2'],
                        f"{corr['corr']:.3f}",
                        strength
                    ])
                
                corr_table = ReportTable(
                    id="T5-相关性分析",
                    title="指标相关性矩阵",
                    headers=["指标A", "指标B", "相关系数", "相关强度"],
                    rows=corr_data,
                    description="各数值指标之间的相关性分析",
                    metrics={"strong_correlations": len([c for c in correlations if abs(c['corr']) > 0.7])}
                )
                tables.append(corr_table)
                
                # 最强相关
                strongest = correlations[0]
                insights.append(HierarchicalInsight(
                    level="微观",
                    dimension="关联性分析",
                    title="指标联动效应",
                    insight=f"{strongest['col1']}与{strongest['col2']}存在{'强' if abs(strongest['corr']) > 0.7 else '中等'}相关（r={strongest['corr']:.3f}）",
                    data_support=f"分析了{len(self.numeric_cols)}个指标间的{len(correlations)}对相关性",
                    values={
                        "指标A": strongest['col1'],
                        "指标B": strongest['col2'],
                        "相关系数": f"{strongest['corr']:.3f}",
                        "相关对数": len(correlations),
                        "强相关对数": len([c for c in correlations if abs(c['corr']) > 0.7])
                    },
                    table_refs=["T5-相关性分析"],
                    recommendation="建议综合考虑相关指标的联动影响，避免单一指标决策"
                ))
        
        # 洞察6: 异常值检测（微观数据质量）
        anomalies = []
        for col in self.numeric_cols[:3]:
            Q1 = self.df[col].quantile(0.25)
            Q3 = self.df[col].quantile(0.75)
            IQR = Q3 - Q1
            lower = Q1 - 1.5 * IQR
            upper = Q3 + 1.5 * IQR
            
            outliers = self.df[(self.df[col] < lower) | (self.df[col] > upper)]
            if len(outliers) > 0:
                anomalies.append({
                    'column': col,
                    'count': len(outliers),
                    'percentage': len(outliers) / len(self.df) * 100,
                    'lower': lower,
                    'upper': upper
                })
        
        if anomalies:
            anomaly_data = []
            for ano in anomalies:
                anomaly_data.append([
                    ano['column'],
                    ano['count'],
                    f"{ano['percentage']:.1f}%",
                    f"{ano['lower']:.2f}",
                    f"{ano['upper']:.2f}"
                ])
            
            anomaly_table = ReportTable(
                id="T6-异常值检测",
                title="数据异常值分析",
                headers=["指标", "异常数量", "异常占比", "下限", "上限"],
                rows=anomaly_data,
                description="基于IQR方法检测的异常值统计",
                metrics={"total_anomalies": sum([a['count'] for a in anomalies])}
            )
            tables.append(anomaly_table)
            
            total_anomalies = sum([a['count'] for a in anomalies])
            insights.append(HierarchicalInsight(
                level="微观",
                dimension="数据质量",
                title="异常值识别",
                insight=f"发现{len(anomalies)}个指标存在异常值，共计{total_anomalies}条记录",
                data_support="基于IQR统计方法检测",
                values={
                    "异常指标数": len(anomalies),
                    "异常记录数": total_anomalies,
                    "异常占比": f"{total_anomalies/len(self.df)*100:.1f}%",
                    "检测字段数": len(self.numeric_cols[:3])
                },
                table_refs=["T6-异常值检测"],
                recommendation="建议核查异常值原因，区分真实业务异常与数据质量问题"
            ))
        
        # 洞察7: 细分项详细分析（最微观）
        if self.name_col and self.numeric_cols:
            primary_metric = self.numeric_cols[0]
            
            # 计算每个细分项的指标
            detail_data = []
            for _, row in self.df.iterrows():
                detail_data.append([
                    row[self.name_col],
                    self.category_col and row.get(self.category_col, 'N/A') or 'N/A',
                    f"{row[primary_metric]:.2f}",
                    f"{(row[primary_metric] / self.df[primary_metric].sum() * 100):.1f}%"
                ])
            
            # 按指标排序
            detail_data.sort(key=lambda x: float(x[2]), reverse=True)
            
            detail_table = ReportTable(
                id="T7-细分项详情",
                title=f"各{self.name_col}{primary_metric}明细",
                headers=[self.name_col, self.category_col or "类别", primary_metric, "占比"],
                rows=detail_data[:15],  # 前15名
                description=f"所有{self.name_col}的{primary_metric}详细数据",
                metrics={"total_items": len(detail_data)}
            )
            tables.append(detail_table)
            
            avg_value = self.df[primary_metric].mean()
            above_avg = len(self.df[self.df[primary_metric] > avg_value])
            
            insights.append(HierarchicalInsight(
                level="微观",
                dimension="细分项分析",
                title="个体表现分布",
                insight=f"共有{above_avg}个{self.name_col}的{primary_metric}高于平均值({avg_value:.2f})",
                data_support=f"平均值{avg_value:.2f}，中位数{self.df[primary_metric].median():.2f}",
                values={
                    "总项数": len(self.df),
                    "高于平均": above_avg,
                    "低于平均": len(self.df) - above_avg,
                    "平均值": f"{avg_value:.2f}",
                    "中位数": f"{self.df[primary_metric].median():.2f}"
                },
                table_refs=["T7-细分项详情"],
                recommendation="建议对高于平均的个体进行经验总结，对低于平均的个体进行帮扶"
            ))
        
        return insights, tables


def generate_hierarchical_report(
    df: pd.DataFrame,
    title: str = "数据分析报告",
    sheet_name: str = "数据表"
) -> Dict[str, Any]:
    """
    生成层级化分析报告
    """
    analyzer = HierarchicalAnalyzer(df)
    
    # 三个层级的分析
    macro_insights, macro_tables = analyzer.analyze_macro_level()
    meso_insights, meso_tables = analyzer.analyze_meso_level()
    micro_insights, micro_tables = analyzer.analyze_micro_level()
    
    all_insights = macro_insights + meso_insights + micro_insights
    all_tables = macro_tables + meso_tables + micro_tables
    
    # 生成Markdown报告
    md_content = _generate_hierarchical_markdown(
        df, title, sheet_name, all_insights, all_tables
    )
    
    # 生成Excel报告
    excel_buffer = _generate_hierarchical_excel(df, all_tables)
    
    # 生成Word报告（使用原有的JS方法）
    from professional_report_integration import generate_professional_report, ProfessionalKeyInsight
    
    key_insights = [
        ProfessionalKeyInsight(
            insight=ins.insight,
            data_support=ins.data_support,
            values=ins.values
        )
        for ins in all_insights[:8]
    ]
    
    word_buffer = generate_professional_report(
        df=df,
        title=title,
        sheet_name=sheet_name,
        key_insights=key_insights,
        summary=f"本报告基于{sheet_name}的清洗后数据生成，共{len(df)}条记录。报告按照宏观-中观-微观三个层级进行系统化分析，涵盖总体核心指标、分类维度分解和指标关联性分析。"
    )
    
    return {
        'markdown': md_content,
        'word': word_buffer,
        'excel': excel_buffer.getvalue() if excel_buffer else None,
        'insights': all_insights,
        'tables': all_tables,
        'structure': {
            'macro': {'insights': len(macro_insights), 'tables': len(macro_tables)},
            'meso': {'insights': len(meso_insights), 'tables': len(meso_tables)},
            'micro': {'insights': len(micro_insights), 'tables': len(micro_tables)}
        }
    }


def _generate_hierarchical_markdown(
    df: pd.DataFrame,
    title: str,
    sheet_name: str,
    insights: List[HierarchicalInsight],
    tables: List[ReportTable]
) -> str:
    """生成层级化Markdown报告"""
    lines = [
        f"# {title}",
        "",
        f"**数据表**: {sheet_name}",
        f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**数据规模**: {len(df)}条记录，{len(df.columns)}个字段",
        "",
        "## 执行摘要",
        "",
        f"本报告基于{sheet_name}的清洗后数据进行系统化层级分析。",
        f"数据集包含{len(df)}条记录，{len(df.columns)}个字段。",
        "报告按照**宏观-中观-微观**三个层级展开，确保分析的系统性和完整性。",
        "",
        "---",
        "",
        "## 第一部分：宏观层分析 - 总体核心指标",
        "",
        "> 从全局视角把握数据总体规模与核心指标表现",
        ""
    ]
    
    # 宏观层
    macro_insights = [ins for ins in insights if ins.level == "宏观"]
    for i, ins in enumerate(macro_insights, 1):
        lines.extend(_format_insight(ins, i))
    
    lines.extend([
        "",
        "---",
        "",
        "## 第二部分：中观层分析 - 指标层级分解",
        "",
        "> 按维度分解指标，从总体到分类逐级深入",
        ""
    ])
    
    # 中观层
    meso_insights = [ins for ins in insights if ins.level == "中观"]
    for i, ins in enumerate(meso_insights, 1):
        lines.extend(_format_insight(ins, i))
    
    lines.extend([
        "",
        "---",
        "",
        "## 第三部分：微观层分析 - 关联性与机制",
        "",
        "> 深入分析指标间关联关系与数据质量",
        ""
    ])
    
    # 微观层
    micro_insights = [ins for ins in insights if ins.level == "微观"]
    for i, ins in enumerate(micro_insights, 1):
        lines.extend(_format_insight(ins, i))
    
    # 数据表格部分
    lines.extend([
        "",
        "---",
        "",
        "## 数据支撑表格",
        "",
        "> 以下表格为上述分析提供详细数据支撑",
        ""
    ])
    
    for table in tables:
        lines.extend([
            f"### {table.title}",
            "",
            f"**表格ID**: {table.id}",
            "",
            f"*{table.description}*",
            "",
            "| " + " | ".join(table.headers) + " |",
            "| " + " | ".join(["---"] * len(table.headers)) + " |"
        ])
        
        for row in table.rows[:20]:
            row_values = [str(v) if v is not None else "" for v in row]
            lines.append("| " + " | ".join(row_values) + " |")
        
        if len(table.rows) > 20:
            lines.append(f"\n*... 共 {len(table.rows)} 行数据 ...*")
        
        lines.append("")
    
    # 结论
    lines.extend([
        "",
        "---",
        "",
        "## 结论与建议",
        "",
        "基于以上三个层级的系统化分析，提出以下建议：",
        "",
        "### 宏观层面",
        "- 持续监控总体核心指标的变化趋势",
        "- 建立指标预警机制，及时发现异常波动",
        "",
        "### 中观层面",
        "- 优化分类维度的资源配置",
        "- 缩小类别间差距，促进均衡发展",
        "",
        "### 微观层面",
        "- 关注指标间的联动效应",
        "- 提升数据质量，确保分析准确性",
        "",
        "---",
        "",
        "*本报告采用系统化层级分析方法生成，数据基于清洗后的原始数据。*"
    ])
    
    return "\n".join(lines)


def _format_insight(ins: HierarchicalInsight, index: int) -> List[str]:
    """格式化洞察输出"""
    lines = [
        f"### {index}. {ins.title}",
        "",
        f"**【{ins.dimension}】** {ins.insight}",
        "",
        f"**数据支撑**: {ins.data_support}",
        ""
    ]
    
    if ins.values:
        lines.append("**关键指标**:")
        for key, value in ins.values.items():
            lines.append(f"- {key}: {value}")
        lines.append("")
    
    if ins.table_refs:
        lines.append(f"**支撑表格**: {', '.join(ins.table_refs)}")
        lines.append("")
    
    if ins.recommendation:
        lines.extend([
            f"**建议**: {ins.recommendation}",
            ""
        ])
    
    return lines


def _generate_hierarchical_excel(df: pd.DataFrame, tables: List[ReportTable]) -> Optional[io.BytesIO]:
    """生成层级化Excel报告"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: 原始数据
        ws_data = wb.create_sheet("原始数据")
        headers = df.columns.tolist()
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_data.column_dimensions[column].width = min(max_length + 2, 50)
        
        # 其他表格按层级分组
        macro_tables = [t for t in tables if t.id.startswith('T1') or t.id.startswith('T2')]
        meso_tables = [t for t in tables if t.id.startswith('T3') or t.id.startswith('T4')]
        micro_tables = [t for t in tables if t.id.startswith('T5') or t.id.startswith('T6') or t.id.startswith('T7')]
        
        for group_name, group_tables in [('宏观层', macro_tables), ('中观层', meso_tables), ('微观层', micro_tables)]:
            for table in group_tables:
                ws = wb.create_sheet(f"{group_name}-{table.id}")
                
                # 标题
                ws.cell(row=1, column=1, value=table.title).font = Font(bold=True, size=14, color="2E74B5")
                ws.cell(row=2, column=1, value=table.description).font = Font(italic=True, size=10, color="666666")
                
                # 表头
                for col_idx, header in enumerate(table.headers, 1):
                    cell = ws.cell(row=4, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E74B5", end_color="2E74B5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 数据
                for row_idx, row in enumerate(table.rows, 5):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"生成 Excel 失败: {e}")
        return None
