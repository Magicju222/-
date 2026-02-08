import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()
ws = wb.active

# 1. Noise Title (Row 1)
ws.merge_cells('A1:E1')
ws['A1'] = "2023年度区域销售业绩汇总表"
ws['A1'].alignment = Alignment(horizontal='center')

# 2. Multi-level Headers (Row 3-4)
# Column A: Product (Vertically merged)
ws['A3'] = "产品名称"
ws['A4'] = "产品名称"
ws.merge_cells('A3:A4')

# Columns B-C: North Region (Horizontally merged)
ws['B3'] = "华北区"
ws['C3'] = "华北区"
ws.merge_cells('B3:C3')
ws['B3'].alignment = Alignment(horizontal='center')

# Columns D-E: South Region (Horizontally merged)
ws['D3'] = "华南区"
ws['E3'] = "华南区"
ws.merge_cells('D3:E3')
ws['D3'].alignment = Alignment(horizontal='center')

# Level 2 Headers
ws['B4'] = "第一季度"
ws['C4'] = "第二季度"
ws['D4'] = "第一季度"
ws['E4'] = "第二季度"

# 3. Data (Row 5-7)
# Merged "Mobile Phone" in Column A
ws['A5'] = "手机"
ws['A6'] = "手机"
ws.merge_cells('A5:A6')
ws['A5'].alignment = Alignment(vertical='center')

ws['B5'] = 1000; ws['C5'] = 1200; ws['D5'] = 800; ws['E5'] = 900
ws['B6'] = 1100; ws['C6'] = 1300; ws['D6'] = 850; ws['E6'] = 950

# Unmerged "Laptop"
ws['A7'] = "笔记本"
ws['B7'] = 5000; ws['C7'] = 5200; ws['D7'] = 4800; ws['E7'] = 4900

# 4. Footer Noise (Row 9-10)
ws['A9'] = "备注：未经允许不得外传"
ws['A10'] = "制表日期：2023-10-01"

wb.save('e:\\徐衡文档\\AI\\Trae EXCEL\\demo.xlsx')
print("demo.xlsx created successfully")
